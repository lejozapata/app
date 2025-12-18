# pacientes_excel.py
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from db import (
    crear_paciente, 
    actualizar_paciente, 
    obtener_paciente, 
    listar_pacientes
)

# ----------------------------
# Excepciones y validador
# ----------------------------
class PlantillaExcelInvalidaError(Exception):
    """El archivo Excel no corresponde a la plantilla esperada."""


class ImportExcelError(Exception):
    """Error general controlado durante la importación."""


# ----------------------------
# Configuración de plantilla
# ----------------------------
SHEET_NAME = "pacientes"

# Orden EXACTO de columnas del Excel (plantilla)
CAMPOS = [
    "documento",
    "tipo_documento",
    "nombre_completo",
    "fecha_nacimiento",  # DD-MM-YYYY
    "sexo",
    "estado_civil",
    "escolaridad",
    "eps",
    "direccion",
    "email",
    "indicativo_pais",  # dígitos (ej: 57)
    "telefono",
    "contacto_emergencia_nombre",
    "contacto_emergencia_telefono",
    "observaciones",
]

OBLIGATORIOS = {"documento", "tipo_documento", "nombre_completo", "fecha_nacimiento"}

HEADER_HELP = {
    "fecha_nacimiento": "Formato: DD-MM-YYYY",
    "indicativo_pais": "Ej: 57 (solo dígitos)",
    "telefono": "Solo dígitos (sin +, sin espacios)",
    "contacto_emergencia_telefono": "Solo dígitos",
}

LISTAS_VALIDACION = {
    "tipo_documento": ["CC", "TI", "CE", "PP", "NIT"],
    "sexo": ["Masculino", "Femenino", "Otro", ""],
    "estado_civil": ["Soltero(a)", "Casado(a)", "Unión libre", "Divorciado(a)", "Viudo(a)", ""],
}


@dataclass
class ImportResult:
    insertados: int
    actualizados: int
    omitidos: int
    errores: List[str]      # errores por fila
    warnings: List[str]     # avisos no fatales (ej: filas vacías)
    
def validar_archivo_pacientes_excel(path: str | Path) -> None:
    """
    Valida que el Excel sea importable:
    - exista la hoja SHEET_NAME (o al menos una hoja con headers compatibles)
    - tenga TODAS las columnas requeridas (CAMPOS)
    - headers en fila 1 no vacíos
    Lanza PlantillaExcelInvalidaError si no cumple.
    """
    path = Path(path)

    try:
        wb = load_workbook(path)
    except Exception as ex:
        raise PlantillaExcelInvalidaError(f"No se pudo abrir el archivo. ¿Es un .xlsx válido? Detalle: {ex}")

    # Preferimos la hoja por nombre; si no existe, tomamos active pero validamos igual
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    # Leer headers (fila 1) hasta max_column
    headers = []
    for i in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=i).value
        headers.append("" if h is None else str(h).strip())

    headers_limpios = [h for h in headers if h]
    if not headers_limpios:
        raise PlantillaExcelInvalidaError("La fila 1 no tiene encabezados. Este archivo no parece una plantilla válida.")

    header_map = {h: idx + 1 for idx, h in enumerate(headers) if h}
    faltantes = [c for c in CAMPOS if c not in header_map]
    if faltantes:
        raise PlantillaExcelInvalidaError(
            "El archivo no coincide con la plantilla. "
            f"Faltan columnas: {', '.join(faltantes)}"
        )

    # Validación adicional: al menos 'documento' debe estar en columna A (opcional pero útil)
    col_doc = header_map.get("documento")
    if col_doc != 1:
        # No lo hacemos fatal si no quieres, pero ayuda a detectar plantillas “rotas”
        raise PlantillaExcelInvalidaError(
            "Encabezados alterados: la columna 'documento' debería estar en la columna A (primera columna)."
        )


def _solo_digitos(s: str) -> str:
    return "".join(c for c in (s or "") if c.isdigit())


def _email_valido(correo: str) -> bool:
    if not correo:
        return True
    return re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", correo) is not None


def _parse_fecha_ddmmyyyy(valor: Any) -> Optional[str]:
    """
    Devuelve string DD-MM-YYYY o None si viene vacío.
    Acepta:
      - str 'DD-MM-YYYY'
      - datetime/date (excel real)
    """
    if valor is None or str(valor).strip() == "":
        return None

    if isinstance(valor, datetime):
        return valor.strftime("%d-%m-%Y")

    # openpyxl puede dar datetime.date en algunos casos
    if hasattr(valor, "strftime"):
        try:
            return valor.strftime("%d-%m-%Y")
        except Exception:
            pass

    s = str(valor).strip()
    # Permitimos DD/MM/YYYY -> lo normalizamos
    s = s.replace("/", "-")
    # Validación estricta
    datetime.strptime(s, "%d-%m-%Y")
    return s


def crear_plantilla_pacientes(path: str | Path) -> Path:
    """
    Crea una plantilla Excel vacía con encabezados, validaciones y estilo avanzado.
    - Fila 1: headers (azul, bloqueados)
    - Fila 2: hints/ayuda (gris, bloqueados)
    - Fila 3+: datos (editables)
    - Fecha con formato REAL DD-MM-YYYY
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Protection
    from openpyxl.worksheet.datavalidation import DataValidation

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # =============================
    # Estilos
    # =============================
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    hint_font = Font(italic=True, color="808080")
    hint_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

    locked = Protection(locked=True)
    unlocked = Protection(locked=False)

    # =============================
    # Headers + Hints
    # =============================
    for col_idx, campo in enumerate(CAMPOS, start=1):
        # ---- Fila 1: HEADER (bloqueado) ----
        header_cell = ws.cell(row=1, column=col_idx, value=campo)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.protection = locked

        # ---- Fila 2: HINT (bloqueado) ----
        help_txt = HEADER_HELP.get(campo, "")
        hint_cell = ws.cell(row=2, column=col_idx, value=help_txt)
        hint_cell.font = hint_font
        hint_cell.fill = hint_fill
        hint_cell.protection = locked

        # ---- Ancho de columna ----
        ws.column_dimensions[header_cell.column_letter].width = max(18, len(campo) + 2)

    # Congelar filas 1 y 2
    ws.freeze_panes = "A3"

    # =============================
    # Validaciones tipo lista
    # =============================
    for campo, opciones in LISTAS_VALIDACION.items():
        if campo not in CAMPOS:
            continue
        col = CAMPOS.index(campo) + 1
        formula = '"' + ",".join(opciones) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=3, column=col).coordinate}:{ws.cell(row=5000, column=col).coordinate}")

    # =============================
    # Formato REAL de fecha (DD-MM-YYYY)
    # =============================
    if "fecha_nacimiento" in CAMPOS:
        col_fecha = CAMPOS.index("fecha_nacimiento") + 1
        col_letter = ws.cell(row=1, column=col_fecha).column_letter

        # Formato visual real
        for r in range(3, 5001):
            cell = ws[f"{col_letter}{r}"]
            cell.number_format = "DD-MM-YYYY"
            cell.protection = unlocked

        # Validación de fecha (Excel)
        dv_fecha = DataValidation(
            type="date",
            operator="between",
            formula1="DATE(1900,1,1)",
            formula2="DATE(2100,12,31)",
            allow_blank=True,
            showErrorMessage=True,
            error="Ingrese una fecha válida en formato DD-MM-YYYY",
            errorTitle="Fecha inválida",
        )
        ws.add_data_validation(dv_fecha)
        dv_fecha.add(f"{col_letter}3:{col_letter}5000")

    # =============================
    # Desbloquear filas de datos
    # =============================
    for row in range(3, 5001):
        for col in range(1, len(CAMPOS) + 1):
            ws.cell(row=row, column=col).protection = unlocked

    # =============================
    # Proteger hoja (clave simple)
    # =============================
    ws.protection.enable()
    ws.protection.set_password("sara")  # puedes cambiarla o quitar password

    wb.save(path)
    return path


def exportar_pacientes_a_excel(path: str | Path) -> Path:
    """
    Exporta TODOS los pacientes a un Excel idéntico a la plantilla:
    - genera plantilla desde cero (hints gris, validaciones, fecha, protección, etc.)
    - rellena datos desde fila 3
    - escribe fecha_nacimiento como datetime real para que Excel aplique DD-MM-YYYY
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # 1) SIEMPRE regenerar la plantilla en el destino (garantiza mismo formato/validaciones)
    crear_plantilla_pacientes(path)

    # 2) Abrir la plantilla recién creada y llenar datos
    wb = load_workbook(path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    pacientes = [dict(p) for p in listar_pacientes()]

    row = 3
    for p in pacientes:
        for col_idx, campo in enumerate(CAMPOS, start=1):
            val = p.get(campo, "")

            # normalizaciones
            if campo in ("telefono", "contacto_emergencia_telefono", "indicativo_pais"):
                val = _solo_digitos(str(val))

            # fecha como datetime real (si viene como DD-MM-YYYY)
            if campo == "fecha_nacimiento":
                s = ("" if val is None else str(val).strip()).replace("/", "-")
                if s:
                    try:
                        val = datetime.strptime(s, "%d-%m-%Y")
                    except Exception:
                        # si por alguna razón viene mal, lo dejamos como texto para que el import lo reporte
                        val = s

            ws.cell(row=row, column=col_idx, value=val if val is not None else "")

        row += 1

    wb.save(path)
    return path


def importar_pacientes_desde_excel(path: str | Path) -> ImportResult:
    path = Path(path)

    # 1) Validación “dura” de plantilla
    validar_archivo_pacientes_excel(path)

    wb = load_workbook(path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    headers = [str(ws.cell(row=1, column=i).value or "").strip() for i in range(1, ws.max_column + 1)]
    header_map = {h: idx + 1 for idx, h in enumerate(headers) if h}

    insertados = 0
    actualizados = 0
    omitidos = 0
    errores: List[str] = []
    warnings: List[str] = []

    # Si el archivo está “vacío” (sin data)
    if ws.max_row < 3:
        warnings.append("El archivo no tiene filas de datos (empiezan en la fila 3). No se importó nada.")
        return ImportResult(insertados, actualizados, omitidos, errores, warnings)

    for r in range(3, ws.max_row + 1):
        fila: Dict[str, Any] = {}
        for campo in CAMPOS:
            c = header_map[campo]
            fila[campo] = ws.cell(row=r, column=c).value

        if all((v is None or str(v).strip() == "") for v in fila.values()):
            omitidos += 1
            continue

        try:
            paciente: Dict[str, str] = {}
            for k, v in fila.items():
                paciente[k] = ("" if v is None else str(v).strip())

            # Obligatorios
            for req in OBLIGATORIOS:
                if not paciente.get(req):
                    raise ValueError(f"Campo obligatorio vacío: {req}")

            # Fecha
            fn = _parse_fecha_ddmmyyyy(fila.get("fecha_nacimiento"))
            if not fn:
                raise ValueError("Campo obligatorio vacío: fecha_nacimiento")
            paciente["fecha_nacimiento"] = fn

            # Email
            if not _email_valido(paciente.get("email", "")):
                raise ValueError("Email inválido")

            # Dígitos
            paciente["indicativo_pais"] = _solo_digitos(paciente.get("indicativo_pais", "")) or "57"
            paciente["telefono"] = _solo_digitos(paciente.get("telefono", ""))
            paciente["contacto_emergencia_telefono"] = _solo_digitos(paciente.get("contacto_emergencia_telefono", ""))

            # Upsert
            existe = obtener_paciente(paciente["documento"])
            if existe:
                actualizar_paciente(paciente)
                actualizados += 1
            else:
                crear_paciente(paciente)
                insertados += 1

        except Exception as ex:
            # Error por fila, no tumba toda la importación
            errores.append(f"Fila {r} (doc: {paciente.get('documento','')}): {ex}")

    if omitidos > 0:
        warnings.append(f"Se omitieron {omitidos} filas vacías.")

    return ImportResult(insertados, actualizados, omitidos, errores, warnings)
